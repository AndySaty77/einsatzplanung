#!/usr/bin/env python3
"""
Import Excel Einsatzplanung_V1.xlsx into Supabase.
Reads the 2026 sheet and creates:
  - Missing employees (inactive/ausgeschieden ones)
  - Projects (ep_projekte)
  - Einplanungen (weekly project assignments)
  - Abwesenheiten (absences)
"""

import json
import requests
import openpyxl
from datetime import date, timedelta

SUPABASE_URL = "https://dnzcjfdrlkuhioyxwgzx.supabase.co"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRuemNqZmRybGt1aGlveXh3Z3p4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzQ0MDk3MCwiZXhwIjoyMDkzMDE2OTcwfQ"
    ".UuwffExao0tWQK-b7Otyvvx8fK_drOEX2jKgpmvOLB8"
)
EXCEL_PATH = "/root/import/Einsatzplanung_V1.xlsx"
YEAR = 2026

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def api(method, path, **kwargs):
    r = requests.request(method, f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS, **kwargs)
    if not r.ok:
        raise RuntimeError(f"{method} {path} → {r.status_code}: {r.text}")
    return r.json() if r.text else []


def monday_of_isoweek(year, week):
    return date.fromisocalendar(year, week, 1)


# ── Absence classification ────────────────────────────────────

ABWESENHEIT_MAP = {
    "urlaub": "Urlaub",
    "elternzeit": "Elternzeit",
    "schule": "Schule",
    "ausgeschieden": "Ausgeschieden",
    "krank": "Krank",
}

SKIP_VALUES = {"kd", "lager", "mit role", "verschiedene", "ausbildung", ""}

PROJECT_ALIASES = {
    "mit mirco": "Breissacher Hof HZ",
    "teningen l.j. str": "Teningen L.J. Str.",
    "waldorfschule ": "Waldorfschule",
}

PROJECT_COLORS = [
    "#3b82f6", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6",
    "#06b6d4", "#84cc16", "#f97316", "#ec4899", "#14b8a6",
    "#6366f1", "#a3e635", "#fb923c", "#e879f9", "#38bdf8",
    "#4ade80", "#fbbf24", "#f87171", "#a78bfa", "#34d399",
]


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Einsatzplanung 2026"]

    # KW → 0-based column index (the Monday column)
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    kw_cols = {}
    for i, val in enumerate(row2):
        if val and str(val).startswith("KW"):
            kw_num = int(str(val).replace("KW", ""))
            kw_cols[kw_num] = i

    employees = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0]
        rolle = row[1]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        rolle = rolle.strip() if rolle else ""

        # Skip Kundendienst and Elektriker — not managed in this tool
        if rolle in ("Kundendienst", "Elektriker"):
            continue

        assignments = {}
        for kw, col in kw_cols.items():
            val = row[col]
            if val and isinstance(val, str):
                val = val.strip()
            if val:
                assignments[kw] = val

        employees.append({"name": name, "rolle": rolle, "assignments": assignments})

    return employees


def main():
    print("── 1. Parsing Excel ────────────────────────────────")
    emp_data = parse_excel()
    for e in emp_data:
        vals = sorted(set(e["assignments"].values()))
        print(f"  {e['name']:25s} {e['rolle']:14s}  {len(e['assignments']):3d} KWs  {vals}")

    # ── 2. Fetch existing employees ──────────────────────────
    print("\n── 2. Employees in Supabase ────────────────────────")
    existing_ma = api("GET", "mitarbeiter?select=id,name,rolle,aktiv")
    ma_by_name = {m["name"]: m for m in existing_ma}
    print(f"  {len(existing_ma)} employees found")

    # Name aliases: Excel name → DB name (defined early, needed for step 3)
    name_alias = {
        "Luka": "Luka Shavana",
    }

    # ── 3. Add missing employees (ausgeschieden ones) ────────
    print("\n── 3. Adding missing employees ─────────────────────")
    for e in emp_data:
        # Resolve alias first — if the alias name exists, no need to add
        resolved_name = name_alias.get(e["name"], e["name"])
        if resolved_name not in ma_by_name and e["name"] not in ma_by_name:
            # Mark as inactive (ausgeschieden)
            new_ma = api("POST", "mitarbeiter", json={
                "name": e["name"],
                "rolle": e["rolle"],
                "aktiv": False,
            })
            if isinstance(new_ma, list):
                new_ma = new_ma[0]
            ma_by_name[e["name"]] = new_ma
            print(f"  ✚ Added (inactive): {e['name']} ({e['rolle']})")

    # ── 4. Build name → id map ───────────────────────────────
    print("\n── 4. Building name map ────────────────────────────")
    # Leiharbeiter: map first excel occurrence to first DB id, second to second
    leiharbeiter_db_ids = [m["id"] for m in existing_ma if m["name"] == "Leiharbeiter"]
    leiharbeiter_excel_count = 0

    def get_ma_id(excel_name):
        nonlocal leiharbeiter_excel_count
        if excel_name == "Leiharbeiter":
            idx = leiharbeiter_excel_count % len(leiharbeiter_db_ids)
            leiharbeiter_excel_count += 1
            return leiharbeiter_db_ids[idx] if leiharbeiter_db_ids else None
        resolved = name_alias.get(excel_name, excel_name)
        m = ma_by_name.get(resolved)
        return m["id"] if m else None

    for e in emp_data:
        mid = get_ma_id(e["name"])
        print(f"  {e['name']:25s} → {mid or '❌ NOT FOUND'}")

    # ── 5. Collect all distinct project names ────────────────
    print("\n── 5. Collecting project names ─────────────────────")
    all_values = set()
    for e in emp_data:
        all_values.update(v.strip() for v in e["assignments"].values())

    project_names = set()
    for v in all_values:
        vl = v.lower()
        if vl in ABWESENHEIT_MAP:
            continue
        if vl in SKIP_VALUES:
            continue
        # Apply aliases to canonicalize
        canonical = PROJECT_ALIASES.get(vl, v)
        project_names.add(canonical)

    project_names = sorted(project_names)
    print(f"  Projects to import: {project_names}")

    # ── 6. Fetch existing projects ───────────────────────────
    print("\n── 6. Fetching existing projects ───────────────────")
    existing_proj = api("GET", "ep_projekte?select=id,name")
    proj_by_name = {p["name"]: p for p in existing_proj}
    print(f"  {len(existing_proj)} existing projects")

    # ── 7. Create missing projects ───────────────────────────
    print("\n── 7. Creating missing projects ────────────────────")
    color_idx = 0
    for pname in project_names:
        if pname in proj_by_name:
            print(f"  ✓ Already exists: {pname}")
            continue
        # Pick a color not yet used
        color = PROJECT_COLORS[color_idx % len(PROJECT_COLORS)]
        color_idx += 1
        new_proj = api("POST", "ep_projekte", json={
            "name": pname,
            "status": "aktiv",
            "farbe": color,
        })
        if isinstance(new_proj, list):
            new_proj = new_proj[0]
        proj_by_name[pname] = new_proj
        print(f"  ✚ Created: {pname} ({color})")

    def resolve_project(raw_val):
        vl = raw_val.lower()
        canonical = PROJECT_ALIASES.get(vl, raw_val)
        return proj_by_name.get(canonical)

    # ── 8. Delete existing 2026 data ─────────────────────────
    print("\n── 8. Clearing existing 2026 einplanungen ──────────")
    kw5_start = monday_of_isoweek(YEAR, 5).isoformat()
    kw52_start = monday_of_isoweek(YEAR, 52).isoformat()
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/einplanungen"
        f"?woche_start=gte.{kw5_start}&woche_start=lte.{kw52_start}",
        headers=HEADERS,
    )
    print(f"  Deleted existing einplanungen: {r.status_code}")

    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/abwesenheiten"
        f"?woche_start=gte.{kw5_start}&woche_start=lte.{kw52_start}",
        headers=HEADERS,
    )
    print(f"  Deleted existing abwesenheiten: {r.status_code}")

    # ── 9. Build and insert einplanungen + abwesenheiten ─────
    print("\n── 9. Inserting assignments ────────────────────────")
    einplanungen_batch = []
    abwesenheiten_batch = []
    skipped = []

    leiharbeiter_excel_count = 0  # reset counter

    for e in emp_data:
        ma_id = get_ma_id(e["name"])
        if not ma_id:
            print(f"  ⚠️  No ID for {e['name']}, skipping")
            continue

        for kw, raw_val in e["assignments"].items():
            raw_stripped = raw_val.strip()
            vl = raw_stripped.lower()
            week_date = monday_of_isoweek(YEAR, kw).isoformat()

            if vl in ABWESENHEIT_MAP:
                abwesenheiten_batch.append({
                    "mitarbeiter_id": ma_id,
                    "woche_start": week_date,
                    "woche_ende": week_date,
                    "typ": ABWESENHEIT_MAP[vl],
                    "notiz": None,
                })
            elif vl in SKIP_VALUES:
                skipped.append(f"{e['name']} KW{kw}: {raw_stripped}")
            else:
                proj = resolve_project(raw_stripped)
                if not proj:
                    skipped.append(f"{e['name']} KW{kw}: {raw_stripped} (no project found)")
                    continue
                einplanungen_batch.append({
                    "mitarbeiter_id": ma_id,
                    "projekt_id": proj["id"],
                    "woche_start": week_date,
                    "bestaetigt": True,
                    "notiz": None,
                })

    print(f"  Einplanungen to insert: {len(einplanungen_batch)}")
    print(f"  Abwesenheiten to insert: {len(abwesenheiten_batch)}")
    print(f"  Skipped entries: {len(skipped)}")
    for s in skipped:
        print(f"    – {s}")

    # Insert in batches of 100
    def batch_insert(table, rows, batch_size=100):
        total = 0
        for i in range(0, len(rows), batch_size):
            chunk = rows[i:i + batch_size]
            api("POST", f"{table}?on_conflict=mitarbeiter_id,woche_start", json=chunk)
            total += len(chunk)
        return total

    if einplanungen_batch:
        n = batch_insert("einplanungen", einplanungen_batch)
        print(f"  ✔ Inserted {n} einplanungen")

    if abwesenheiten_batch:
        # Abwesenheiten have a different unique constraint — use simple insert
        n = batch_insert("abwesenheiten", abwesenheiten_batch)
        print(f"  ✔ Inserted {n} abwesenheiten")

    print("\n── Done! ───────────────────────────────────────────")


if __name__ == "__main__":
    main()
