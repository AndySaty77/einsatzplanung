#!/usr/bin/env python3
"""
Import Kundendienst + Elektriker employees and their weekly KD assignments.
Creates a 'Kundendienst' project and fills all KD weeks for each employee.
"""

import requests
import openpyxl
from datetime import date

SUPABASE_URL = "https://dnzcjfdrlkuhioyxwgzx.supabase.co"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRuemNqZmRybGt1aGlveXh3Z3p4Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzQ0MDk3MCwiZXhwIjoyMDkzMDE2OTcwfQ"
    ".UuwffExao0tWQK-b7Otyvvx8fK_drOEX2jKgpmvOLB8"
)
EXCEL_PATH = "/root/import/Einsatzplanung_V1.xlsx"
YEAR = 2026

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def api(method, path, **kwargs):
    r = requests.request(method, f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS, **kwargs)
    if not r.ok:
        raise RuntimeError(f"{method} {path} → {r.status_code}: {r.text}")
    return r.json() if r.text else []


def monday_of_isoweek(year, week):
    return date.fromisocalendar(year, week, 1)


def main():
    print("── 1. Excel parsen ─────────────────────────────────")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Einsatzplanung 2026"]

    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    kw_cols = {}
    for i, val in enumerate(row2):
        if val and str(val).startswith("KW"):
            kw_cols[int(str(val).replace("KW", ""))] = i

    kd_employees = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0]
        rolle = row[1]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        rolle = rolle.strip() if rolle else ""
        if rolle not in ("Kundendienst", "Elektriker"):
            continue
        kws = [kw for kw, col in kw_cols.items()
               if row[col] and str(row[col]).strip().upper() == "KD"]
        kd_employees.append({"name": name, "rolle": rolle, "kws": sorted(kws)})
        print(f"  {name:25s} | {rolle:14s} | {len(kws)} KW-Wochen")

    # ── 2. Kundendienst-Projekt anlegen ───────────────────────
    print("\n── 2. Kundendienst-Projekt ─────────────────────────")
    existing_proj = api("GET", "ep_projekte?select=id,name")
    proj_by_name = {p["name"]: p for p in existing_proj}

    if "Kundendienst" not in proj_by_name:
        proj = api("POST", "ep_projekte", json={
            "name": "Kundendienst",
            "status": "aktiv",
            "farbe": "#d97706",
            "ist_arbeitsvorrat": False,
        })
        if isinstance(proj, list):
            proj = proj[0]
        proj_by_name["Kundendienst"] = proj
        print("  ✚ Projekt 'Kundendienst' erstellt")
    else:
        print("  ✓ Projekt 'Kundendienst' bereits vorhanden")

    kd_projekt_id = proj_by_name["Kundendienst"]["id"]

    # ── 3. Mitarbeiter anlegen / prüfen ───────────────────────
    print("\n── 3. Mitarbeiter ──────────────────────────────────")
    existing_ma = api("GET", "mitarbeiter?select=id,name,rolle,aktiv")
    ma_by_name = {m["name"]: m for m in existing_ma}

    for e in kd_employees:
        if e["name"] not in ma_by_name:
            new_ma = api("POST", "mitarbeiter", json={
                "name": e["name"],
                "rolle": e["rolle"],
                "aktiv": True,
            })
            if isinstance(new_ma, list):
                new_ma = new_ma[0]
            ma_by_name[e["name"]] = new_ma
            print(f"  ✚ Angelegt: {e['name']} ({e['rolle']})")
        else:
            print(f"  ✓ Vorhanden: {e['name']}")

    # ── 4. Bestehende Einplanungen dieser Mitarbeiter löschen ─
    print("\n── 4. Alte Einplanungen löschen ────────────────────")
    kw5_start  = monday_of_isoweek(YEAR, 5).isoformat()
    kw52_start = monday_of_isoweek(YEAR, 52).isoformat()
    for e in kd_employees:
        ma_id = ma_by_name[e["name"]]["id"]
        r = requests.delete(
            f"{SUPABASE_URL}/rest/v1/einplanungen"
            f"?mitarbeiter_id=eq.{ma_id}"
            f"&woche_start=gte.{kw5_start}&woche_start=lte.{kw52_start}",
            headers=HEADERS,
        )
        print(f"  {e['name']}: {r.status_code}")

    # ── 5. Einplanungen eintragen ─────────────────────────────
    print("\n── 5. Einplanungen eintragen ───────────────────────")
    batch = []
    for e in kd_employees:
        ma_id = ma_by_name[e["name"]]["id"]
        for kw in e["kws"]:
            batch.append({
                "mitarbeiter_id": ma_id,
                "projekt_id": kd_projekt_id,
                "woche_start": monday_of_isoweek(YEAR, kw).isoformat(),
                "bestaetigt": True,
                "notiz": None,
            })

    print(f"  {len(batch)} Einplanungen")
    for i in range(0, len(batch), 100):
        api("POST", "einplanungen?on_conflict=mitarbeiter_id,woche_start", json=batch[i:i+100])

    print(f"  ✔ Fertig!")
    print("\n── Done! ───────────────────────────────────────────")


if __name__ == "__main__":
    main()
